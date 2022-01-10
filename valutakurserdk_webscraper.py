import requests
from bs4 import BeautifulSoup
import openpyxl

# Get data from valutakurser.dk ----------------------------start-------------------------------

url = "https://www.valutakurser.dk/"

response = requests.get(url)

soup = BeautifulSoup(response.text, 'html.parser')

valutaer = soup.findAll('div', attrs={'class': 'currencyItem_currencyItemWrapper__2-TKC'})

# Get data from valutakurser.dk -------------------------------end------------------------------

# Sort data ------------------------------------------------start-------------------------------
listWithDicts = []
for item in valutaer:
	nameDivs = item.findAll('div', attrs={'class': 'currencyItem_currencyNameContainer__19YHn'})
	valueDivs = item.findAll('div', attrs={'class': 'currencyItem_actualValueContainer__2xLkB'})
	
	nameList = []
	valueList = []

	for x in nameDivs:
		nameList.append(str(x))

	for x in valueDivs:
		valueList.append(str(x))

	nameString = ""
	valueString = ""
	
	for i in nameList:
		nameString = nameString + i

	for i in valueList:
		valueString = valueString + i

	name = nameString.split(">")[1].split("</div")[0]
	value = valueString.split(">")[1].split("</div")[0]

	if name == "Euro" or name == "Britiske pund" or name == "Amerikanske dollar":
		keyValue = [name, value]
		listWithDicts.append(keyValue)

print("\nData er blevet hentet!\n")

# Sort data ------------------------------------------------end---------------------------------

# Manipulate Excel file -------------------------------- start --------------------------------
wb = openpyxl.load_workbook('fil.xlsx')

arkToBeEdited = int(input("Ark nummer som skal ændres: "))
ws = wb.worksheets[arkToBeEdited - 1]

row = 4
finalRow = row + 3

while row < finalRow:
	if ws.cell(row=row, column=12).value == "USD":
		for item in listWithDicts:
			if item[0] == "Amerikanske dollar":
				ws.cell(row=row, column=13).value = item[1]

	if ws.cell(row=row, column=12).value == "GBP":
		for item in listWithDicts:
			if item[0] == "Britiske pund":
				ws.cell(row=row, column=13).value = item[1]

	if ws.cell(row=row, column=12).value == "EUR":
		for item in listWithDicts:
			if item[0] == "Euro":
				ws.cell(row=row, column=13).value = item[1]
	row += 1
		
print("Data indsat!\n")

wb.save('fil.xlsx')

print("Fil gemt!\n")

# Manipulate Excel file -------------------------------- end --------------------------------